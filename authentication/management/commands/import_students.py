from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from authentication.models import UserProfile
from openpyxl import load_workbook
import os


class Command(BaseCommand):
    help = 'Import students from Excel file (Daftar-Siswa-Cleaned.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--filepath',
            type=str,
            help='Path to Excel file',
            default='static/files/Daftar-Siswa-Cleaned.xlsx'
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            help='Skip existing users',
            default=True
        )

    def handle(self, *args, **options):
        filepath = options['filepath']
        skip_existing = options['skip_existing']
        
        # Check if file exists
        if not os.path.exists(filepath):
            raise CommandError(f'File not found: {filepath}')
        
        try:
            workbook = load_workbook(filepath)
            worksheet = workbook.active
            
            imported_count = 0
            skipped_count = 0
            error_count = 0
            
            # Start from row 2 (skip header)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    nis, nama, jenis_kelamin, kelas = row[0], row[1], row[2], row[3]
                    
                    # Validate data
                    if not nis or not nama:
                        self.stdout.write(self.style.WARNING(f'Row {row_idx}: Missing NIS or Nama, skipping'))
                        skipped_count += 1
                        continue
                    
                    # Convert to string and strip whitespace
                    # Handle float format (e.g., 2514440.0 -> 2514440)
                    if isinstance(nis, float):
                        nis = str(int(nis))
                    else:
                        nis = str(nis).strip()
                        # Remove .0 suffix if present
                        if nis.endswith('.0'):
                            nis = nis[:-2]
                    nama = str(nama).strip()
                    jenis_kelamin = str(jenis_kelamin).strip() if jenis_kelamin else ''
                    kelas = str(kelas).strip() if kelas else ''
                    
                    # Check if user already exists
                    if User.objects.filter(username=nis).exists():
                        if skip_existing:
                            self.stdout.write(self.style.WARNING(f'Row {row_idx}: User {nis} already exists, skipping'))
                            skipped_count += 1
                            continue
                        else:
                            self.stdout.write(self.style.ERROR(f'Row {row_idx}: User {nis} already exists'))
                            error_count += 1
                            continue
                    
                    # Parse nama (split into first and last name)
                    parts = nama.split(' ', 1)
                    first_name = parts[0]
                    last_name = parts[1] if len(parts) > 1 else ''
                    
                    # Create user
                    user = User.objects.create_user(
                        username=nis,
                        password=nis,  # Default password is NIS
                        first_name=first_name,
                        last_name=last_name,
                        email=f'{nis}@student.local'
                    )
                    
                    # Create user profile
                    UserProfile.objects.create(
                        user=user,
                        role='student',
                        nis=nis,
                        gender=jenis_kelamin,
                        kelas=kelas
                    )
                    
                    self.stdout.write(self.style.SUCCESS(f'Row {row_idx}: Successfully imported {nis} - {nama}'))
                    imported_count += 1
                    
                except TypeError as e:
                    self.stdout.write(self.style.ERROR(f'Row {row_idx}: Type error - {str(e)}'))
                    error_count += 1
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'Row {row_idx}: Error - {str(e)}'))
                    error_count += 1
            
            # Print summary
            self.stdout.write(self.style.SUCCESS('\n========== SUMMARY =========='))
            self.stdout.write(f'Imported: {imported_count}')
            self.stdout.write(self.style.WARNING(f'Skipped: {skipped_count}'))
            self.stdout.write(self.style.ERROR(f'Errors: {error_count}'))
            self.stdout.write('=============================')
            
        except Exception as e:
            raise CommandError(f'Error reading Excel file: {str(e)}')
