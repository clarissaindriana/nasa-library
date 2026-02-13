from django.apps import AppConfig
import os
import logging

logger = logging.getLogger(__name__)


def import_students_callback(sender, **kwargs):
    """Import students after migrations run"""
    from django.contrib.auth.models import User
    from authentication.models import UserProfile
    from openpyxl import load_workbook
    
    filepath = 'static/files/Daftar-Siswa-Cleaned.xlsx'
    
    # Check if file exists
    if not os.path.exists(filepath):
        logger.warning(f'Student data file not found: {filepath}')
        return
    
    # Check if students have already been imported
    if UserProfile.objects.filter(role='student').exists():
        # Check if any users have the float format issue (NIS with .0)
        problematic_users = User.objects.filter(username__endswith='.0')
        if problematic_users.exists():
            logger.warning(f'Found {problematic_users.count()} users with float-format NIS. Fixing...')
            for user in problematic_users:
                old_username = user.username
                new_username = old_username.replace('.0', '')
                if not User.objects.filter(username=new_username).exists():
                    user.username = new_username
                    user.set_password(new_username)
                    user.save()
                    profile = user.profile
                    profile.nis = new_username
                    profile.save()
                    logger.info(f'Fixed user: {old_username} -> {new_username}')
            logger.info('Float-format NIS fix complete')
        return
    
    try:
        workbook = load_workbook(filepath)
        worksheet = workbook.active
        
        imported_count = 0
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nis, nama, jenis_kelamin, kelas = row[0], row[1], row[2], row[3]
                
                if not nis or not nama:
                    continue
                
                # Handle float format (e.g., 2514440.0 -> 2514440)
                if isinstance(nis, float):
                    nis = str(int(nis))
                else:
                    nis = str(nis).strip()
                    if nis.endswith('.0'):
                        nis = nis[:-2]
                
                nama = str(nama).strip()
                jenis_kelamin = str(jenis_kelamin).strip() if jenis_kelamin else ''
                kelas = str(kelas).strip() if kelas else ''
                
                if User.objects.filter(username=nis).exists():
                    continue
                
                parts = nama.split(' ', 1)
                first_name = parts[0]
                last_name = parts[1] if len(parts) > 1 else ''
                
                user = User.objects.create_user(
                    username=nis,
                    password=nis,
                    first_name=first_name,
                    last_name=last_name,
                    email=f'{nis}@student.local'
                )
                
                UserProfile.objects.create(
                    user=user,
                    role='student',
                    nis=nis,
                    gender=jenis_kelamin,
                    kelas=kelas
                )
                
                imported_count += 1
                
            except Exception as e:
                logger.error(f'Row {row_idx}: Error importing student - {str(e)}')
                continue
        
        if imported_count > 0:
            logger.info(f'Successfully auto-imported {imported_count} students')
        
    except Exception as e:
        logger.error(f'Error reading Excel file: {str(e)}')


class AuthenticationConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'authentication'

    def ready(self):
        # Register the post_migrate signal
        from django.db.models.signals import post_migrate
        post_migrate.connect(import_students_callback, sender=self)
